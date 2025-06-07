# backend/app/api/endpoints/collabcards.py
from __future__ import annotations

import re
import unicodedata
import uuid
from pathlib import Path
from typing import Dict, List

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_user, get_db
from app.core.config import get_settings
from app.models import AuditLog, Batch, CollabCard, User
from app.models.enums import BatchStatus
from app.models.user import Role
from app.schemas.batch import BatchRead
from app.schemas.collabcard import CollabCardCreate, CollabCardRead

# ────────────────────────────────────────────────
# NEW 3. Create batch from existing users
# ────────────────────────────────────────────────
from pydantic import BaseModel, Field, validator
from sqlalchemy import select, func

# ──────────────────────────────────────────────────────────────
# router / constants
# ──────────────────────────────────────────────────────────────
router = APIRouter(prefix="/collabcards", tags=["collabcards"])
settings = get_settings()

UPLOAD_ROOT = Path(settings.upload_dir or "/data/uploads").joinpath("records")


def _collapse(text: str | None) -> str:
    """lower-case, strip accents & punctuation; return '' for None."""
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = "".join(
        c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-z0-9]+", "", text)


SPANISH_TO_ENGLISH: Dict[str, str] = {
    _collapse("Nombre"): "full_name",
    _collapse("Correo Electrónico"): "email",
    _collapse("Correo Electronico"): "email",
    _collapse("Correo"): "email",
    _collapse("Celular"): "mobile_phone",
    _collapse("Puesto"): "job_title",
    _collapse("Teléfono Oficina"): "office_phone",
    _collapse("Teléfono Ofi"): "office_phone",
    _collapse("Telefono Ofi"): "office_phone",
}

HEADER_KEY = _collapse("Nombre")  # first entry → used to spot header rows
STR_FIELDS = {"full_name", "email", "mobile_phone", "job_title", "office_phone"}
REQUIRED_FIELDS = {"full_name", "email", "job_title"}
BATCH_SIZE = 5_000
UPLOAD_CHUNK = 1_048_576  # 1 MiB

# ──────────────────────────────────────────────────────────────
# helpers (unchanged)
# ──────────────────────────────────────────────────────────────
async def _log(
    db: AsyncSession,
    *,
    user_id: uuid.UUID | None,
    entity_type: str,
    entity_id: uuid.UUID,
    action: str,
    details: dict,
) -> None:
    db.add(
        AuditLog(
            user_id=user_id,
            entity_type=entity_type,
            entity_id=entity_id,
            action=action,
            details=details,
        )
    )
    await db.commit()


def _company_guard(current: User, company_id: uuid.UUID | None) -> uuid.UUID | None:
    if current.role == Role.owner:
        return company_id
    if company_id and company_id != current.company_id:
        raise HTTPException(status_code=403, detail="Cannot act on other companies")
    if not current.company_id:
        raise HTTPException(status_code=400, detail="User has no company assigned")
    return current.company_id


# ────────────────────────────────────────────────
# 1. Single record endpoint (unchanged)
# ────────────────────────────────────────────────
@router.post("/", response_model=CollabCardRead, status_code=status.HTTP_201_CREATED)
async def create_single_collabcard(
    body: CollabCardCreate,
    background_tasks: BackgroundTasks,
    company_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    target_company_id = _company_guard(current, company_id)

    batch = Batch(
        company_id=target_company_id,
        created_by=current.id,
        original_filename=None,
        total_records=1,
        processed_records=0,
        status=BatchStatus.pending,
    )
    db.add(batch)
    await db.flush()

    record = CollabCard(
        batch_id=batch.id,
        company_id=target_company_id,
        created_by=current.id,
        **body.model_dump(),
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)

    background_tasks.add_task(
        _log,
        db=db,
        user_id=current.id,
        entity_type="collab_card",
        entity_id=record.id,
        action="create",
        details={"via": "typed"},
    )
    return record


# ────────────────────────────────────────────────
# 2. XLSX upload (streaming)
# ────────────────────────────────────────────────
@router.post(
    "/upload-xlsx",
    response_model=BatchRead,
    status_code=status.HTTP_202_ACCEPTED,
)
async def upload_collabcard_xlsx(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    company_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, detail="File must be .xlsx")

    target_company_id = _company_guard(current, company_id)

    batch = Batch(
        company_id=target_company_id,
        created_by=current.id,
        original_filename=file.filename,
        status=BatchStatus.processing,
    )
    db.add(batch)
    await db.flush()

    # ── save upload ───────────────────────────────────────────
    dest_dir = (
        UPLOAD_ROOT
        / (str(target_company_id) if target_company_id else "global")
        / str(batch.id)
    )
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / file.filename
    with dest_path.open("wb") as out:
        for chunk in iter(lambda: file.file.read(UPLOAD_CHUNK), b""):
            out.write(chunk)

    try:
        wb = load_workbook(dest_path, read_only=True, data_only=True)
        ws = wb.active

        # ── detect header row ────────────────────────────────
        header_map: Dict[int, str] = {}
        for row in ws.iter_rows(values_only=True):
            mapped = [SPANISH_TO_ENGLISH.get(_collapse(c), "") for c in row]
            header_map = {ix: f for ix, f in enumerate(mapped) if f}
            if REQUIRED_FIELDS <= set(header_map.values()):
                break

        if not header_map:
            raise ValueError(f"missing columns {REQUIRED_FIELDS}")

        buffer: List[CollabCard] = []
        total = 0

        # ── row loop ───────────────────────────────────────────
        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue  # blank

            # skip any row whose first mapped cell still contains "nombre"
            if _collapse(row[min(header_map)]) == HEADER_KEY:
                continue

            record = {
                header_map[ix]: val
                for ix, val in enumerate(row)
                if ix in header_map and val not in (None, "")
            }

            if REQUIRED_FIELDS - record.keys():
                continue  # missing mandatory data

            # coerce strings & trim
            for k in STR_FIELDS:
                if k in record and record[k] is not None:
                    record[k] = str(record[k]).strip()

            full_name = record.get("full_name", "")
            if full_name.isdigit():   # row-number row
                continue

            buffer.append(
                CollabCard(
                    batch_id=batch.id,
                    company_id=target_company_id,
                    created_by=current.id,
                    **record,
                )
            )
            total += 1
            if len(buffer) >= BATCH_SIZE:
                db.add_all(buffer)
                buffer.clear()

        if buffer:
            db.add_all(buffer)

        batch.total_records = total
        batch.processed_records = 0
        batch.status = BatchStatus.pending
        await db.commit()

    except Exception as exc:  # noqa: BLE001
        await db.rollback()
        batch.status = BatchStatus.error
        await db.commit()
        raise HTTPException(400, detail=f"Parse error: {exc}") from exc

    background_tasks.add_task(
        _log,
        db=db,
        user_id=current.id,
        entity_type="batch",
        entity_id=batch.id,
        action="upload_xlsx",
        details={"filename": file.filename, "records": batch.total_records},
    )
    return batch

######################## BATCH FROM USERS ##########################
# ------------------------------------------------------------------
# 3A. request / response schemas
# ------------------------------------------------------------------
class _BatchFromUsersBody(BaseModel):
    """Payload for /collabcards/from-users"""

    company_id: uuid.UUID | None = Field(
        None,
        description="Company that owns the batch. "
        "Global owners / admins must supply this. "
        "Company owners / admins must omit or match their own company.",
    )
    user_ids: List[uuid.UUID] = Field(
        ...,
        min_items=1,
        description="List of user IDs to include in the batch",
    )

    # eliminate accidental duplicates early
    @validator("user_ids", pre=True)
    def _dedupe(cls, v):
        return list(dict.fromkeys(v))


# ------------------------------------------------------------------
# 3B. role helper – “administrator” is optional
# ------------------------------------------------------------------
_ALLOWED_ROLES = {"owner", "administrator"}  # string names, tolerant of enum

def _raise_if_not_admin(current: User) -> None:
    if current.role.value not in _ALLOWED_ROLES:  # Role -> str
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only owners or administrators can call this endpoint",
        )


# ------------------------------------------------------------------
# 3C. endpoint implementation
# ------------------------------------------------------------------
@router.post(
    "/from-users",
    response_model=BatchRead,
    status_code=status.HTTP_201_CREATED,
    summary="Create CollabCard batch from existing users",
)
async def create_collabcard_batch_from_users(
    body: _BatchFromUsersBody,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    """
    * Global **owner / administrator** – pass **any** `company_id`
      and *any* users belonging to it.

    * **Company owner / administrator** – `company_id` is optional
      (defaults to caller’s company) **and all users must belong to
      that same company**.
    """
    # ── authorisation ────────────────────────────────────────────
    _raise_if_not_admin(current)  # must be owner/admin

    # resolve company in scope
    target_company_id = _company_guard(current, body.company_id)

    if target_company_id is None:
        raise HTTPException(
            status_code=400,
            detail="company_id is required for a global batch",
        )

    # ── fetch & validate users ──────────────────────────────────
    stmt = (
        select(User)
        .where(User.id.in_(body.user_ids))
        .where(User.company_id == target_company_id)
    )
    users = (await db.execute(stmt)).scalars().all()

    if len(users) != len(body.user_ids):
        found_ids = {u.id for u in users}
        missing = [str(uid) for uid in body.user_ids if uid not in found_ids]
        raise HTTPException(
            status_code=404,
            detail=f"These user_ids were not found for this company: {missing}",
        )

    # ── create batch shell ──────────────────────────────────────
    batch = Batch(
        company_id=target_company_id,
        created_by=current.id,
        original_filename=None,
        total_records=len(users),
        processed_records=0,
        status=BatchStatus.pending,
    )
    db.add(batch)
    await db.flush()  # we need batch.id

    # ── create CollabCard rows in chunks for efficiency ─────────
    buffer: list[CollabCard] = []
    for usr in users:
        buffer.append(
            CollabCard(
                batch_id=batch.id,
                company_id=target_company_id,
                created_by=current.id,
                full_name=usr.card_full_name or usr.email.split("@")[0],
                email=usr.card_email or usr.email,
                mobile_phone=usr.card_mobile_phone,
                job_title=usr.card_job_title,
                office_phone=usr.card_office_phone,
            )
        )
    db.add_all(buffer)
    await db.commit()
    await db.refresh(batch)

    # ── audit entry (fire-and-forget) ───────────────────────────
    background_tasks.add_task(
        _log,
        db=db,
        user_id=current.id,
        entity_type="batch",
        entity_id=batch.id,
        action="create_from_users",
        details={
            "user_count": len(users),
            "company_id": str(target_company_id),
        },
    )

    return batch




######################## GETTING BATCH DATA ##########################
# ────────────────────────────────────────────────
# 4. Utility – role check helper reused below
# ────────────────────────────────────────────────
def _assert_owner_or_admin(user: User) -> None:
    """Allow only global/company *owner* or *administrator*."""
    if user.role.value not in _ALLOWED_ROLES:        # ← defined earlier
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Only owners or administrators can access this endpoint",
        )


# ────────────────────────────────────────────────
# 5. List *in-flight* batches for a company
# ────────────────────────────────────────────────
"""
GET /collabcards/pending-batches
Purpose: All batches for a company created in the last 7 days where processed_records < total_records.
Who can call: Owner / Administrator (global or company).
"""
from datetime import datetime, timedelta
from sqlalchemy import select, and_

@router.get(
    "/pending-batches",
    response_model=list[BatchRead],
    summary="Batches (<7 days) with pending records",
)
async def list_pending_batches(
    company_id: uuid.UUID | None = Query(
        None,
        description="Company to inspect (required for global owners/admins)",
    ),
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    """
    Returns batches **created in the last 7 days** whose
    `processed_records < total_records`.

    * Global owners / administrators – provide any `company_id`.  
    * Company owners / administrators – omit `company_id`
      (defaults to caller’s company).
    """
    _assert_owner_or_admin(current)

    target_company_id = _company_guard(current, company_id)
    if target_company_id is None:
        raise HTTPException(
            400, detail="company_id is required for a global query"
        )

    seven_days_ago = datetime.utcnow() - timedelta(days=7)

    stmt = (
        select(Batch)
        .where(
            and_(
                Batch.company_id == target_company_id,
                Batch.created_at >= seven_days_ago,
                Batch.processed_records < Batch.total_records,
            )
        )
        .order_by(Batch.created_at.desc())
    )
    res = await db.execute(stmt)
    return res.scalars().all()


# ────────────────────────────────────────────────
# 6. Get all CollabCards for a specific batch
# ────────────────────────────────────────────────
@router.get(
    "/batch/{batch_id}/records",
    response_model=list[CollabCardRead],
    summary="All records inside one batch",
)
async def list_collabcards_in_batch(
    batch_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    """
    url: /collabcards/batch/{batch_id}/records
    purpose: Returns all CollabCard rows inside the batch.
    who can: Owner / Administrator of the batch’s company, or any global owner / admin.

    Fetches **every** `CollabCard` that belongs to `batch_id`.

    Access is restricted to owners / administrators
    (global or company-scoped).
    """
    _assert_owner_or_admin(current)

    # ── fetch batch first (authorisation) ──────────────────────
    batch: Batch | None = (
        await db.execute(select(Batch).where(Batch.id == batch_id))
    ).scalar_one_or_none()
    if not batch:
        raise HTTPException(404, detail="Batch not found")

    _company_guard(current, batch.company_id)  # raises if forbidden

    # ── fetch records ─────────────────────────────────────────
    res = await db.execute(
        select(CollabCard)
        .where(CollabCard.batch_id == batch_id)
        .order_by(CollabCard.created_at.asc())
    )
    return res.scalars().all()
