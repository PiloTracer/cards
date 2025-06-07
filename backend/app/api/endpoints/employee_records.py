# backend/app/api/endpoints/employee_records.py
from __future__ import annotations

import re
import unicodedata
import uuid
from pathlib import Path
from typing import Dict, List

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    HTTPException,
    Query,
    UploadFile,
    status,
)
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.dependencies import get_current_user, get_db
from app.core.config import get_settings
from app.models import AuditLog, Batch, EmployeeRecord, User
from app.models.enums import BatchStatus
from app.models.user import Role
from app.schemas.batch import BatchRead
from app.schemas.employee import EmployeeCreate, EmployeeRead

# ──────────────────────────────────────────────────────────────
# router / constants
# ──────────────────────────────────────────────────────────────
router = APIRouter(prefix="/employee-records", tags=["employee_records"])
settings = get_settings()

UPLOAD_ROOT = Path(settings.upload_dir or "/data/uploads").joinpath("records")


def _collapse(text: str | None) -> str:
    """lower-case, strip accents & punctuation; return '' for None."""
    if not isinstance(text, str):
        return ""
    text = text.lower()
    text = "".join(
        c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-z0-9]+", "", text)


SPANISH_TO_ENGLISH: Dict[str, str] = {
    _collapse("Nombre"): "full_name",
    _collapse("Correo Electrónico"): "email",
    _collapse("Correo Electronico"): "email",
    _collapse("Correo"): "email",
    _collapse("Celular"): "mobile_phone",
    _collapse("Puesto"): "job_title",
    _collapse("Teléfono Oficina"): "office_phone",
    _collapse("Teléfono Ofi"): "office_phone",
    _collapse("Telefono Ofi"): "office_phone",
}

HEADER_KEY = _collapse("Nombre")  # first entry → used to spot header rows
STR_FIELDS = {"full_name", "email", "mobile_phone", "job_title", "office_phone"}
REQUIRED_FIELDS = {"full_name", "email", "job_title"}
BATCH_SIZE = 5_000
UPLOAD_CHUNK = 1_048_576  # 1 MiB

# ──────────────────────────────────────────────────────────────
# helpers (unchanged)
# ──────────────────────────────────────────────────────────────
async def _log(
    db: AsyncSession,
    *,
    user_id: uuid.UUID | None,
    entity_type: str,
    entity_id: uuid.UUID,
    action: str,
    details: dict,
) -> None:
    db.add(
        AuditLog(
            user_id=user_id,
            entity_type=entity_type,
            entity_id=entity_id,
            action=action,
            details=details,
        )
    )
    await db.commit()


def _company_guard(current: User, company_id: uuid.UUID | None) -> uuid.UUID | None:
    if current.role == Role.owner:
        return company_id
    if company_id and company_id != current.company_id:
        raise HTTPException(status_code=403, detail="Cannot act on other companies")
    if not current.company_id:
        raise HTTPException(status_code=400, detail="User has no company assigned")
    return current.company_id


# ────────────────────────────────────────────────
# 1. Single record endpoint (unchanged)
# ────────────────────────────────────────────────
@router.post("/", response_model=EmployeeRead, status_code=status.HTTP_201_CREATED)
async def create_single_employee(
    body: EmployeeCreate,
    background_tasks: BackgroundTasks,
    company_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    target_company_id = _company_guard(current, company_id)

    batch = Batch(
        company_id=target_company_id,
        created_by=current.id,
        original_filename=None,
        total_records=1,
        processed_records=0,
        status=BatchStatus.pending,
    )
    db.add(batch)
    await db.flush()

    record = EmployeeRecord(
        batch_id=batch.id,
        company_id=target_company_id,
        created_by=current.id,
        **body.model_dump(),
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)

    background_tasks.add_task(
        _log,
        db=db,
        user_id=current.id,
        entity_type="employee_record",
        entity_id=record.id,
        action="create",
        details={"via": "typed"},
    )
    return record


# ────────────────────────────────────────────────
# 2. XLSX upload (streaming)
# ────────────────────────────────────────────────
@router.post(
    "/upload-xlsx",
    response_model=BatchRead,
    status_code=status.HTTP_202_ACCEPTED,
)
async def upload_employee_xlsx(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    company_id: uuid.UUID | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current: User = Depends(get_current_user),
):
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, detail="File must be .xlsx")

    target_company_id = _company_guard(current, company_id)

    batch = Batch(
        company_id=target_company_id,
        created_by=current.id,
        original_filename=file.filename,
        status=BatchStatus.processing,
    )
    db.add(batch)
    await db.flush()

    # ── save upload ───────────────────────────────────────────
    dest_dir = (
        UPLOAD_ROOT
        / (str(target_company_id) if target_company_id else "global")
        / str(batch.id)
    )
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest_path = dest_dir / file.filename
    with dest_path.open("wb") as out:
        for chunk in iter(lambda: file.file.read(UPLOAD_CHUNK), b""):
            out.write(chunk)

    try:
        wb = load_workbook(dest_path, read_only=True, data_only=True)
        ws = wb.active

        # ── detect header row ────────────────────────────────
        header_map: Dict[int, str] = {}
        for row in ws.iter_rows(values_only=True):
            mapped = [SPANISH_TO_ENGLISH.get(_collapse(c), "") for c in row]
            header_map = {ix: f for ix, f in enumerate(mapped) if f}
            if REQUIRED_FIELDS <= set(header_map.values()):
                break

        if not header_map:
            raise ValueError(f"missing columns {REQUIRED_FIELDS}")

        buffer: List[EmployeeRecord] = []
        total = 0

        # ── row loop ───────────────────────────────────────────
        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue  # blank

            # skip any row whose first mapped cell still contains "nombre"
            if _collapse(row[min(header_map)]) == HEADER_KEY:
                continue

            record = {
                header_map[ix]: val
                for ix, val in enumerate(row)
                if ix in header_map and val not in (None, "")
            }

            if REQUIRED_FIELDS - record.keys():
                continue  # missing mandatory data

            # coerce strings & trim
            for k in STR_FIELDS:
                if k in record and record[k] is not None:
                    record[k] = str(record[k]).strip()

            full_name = record.get("full_name", "")
            if full_name.isdigit():   # row-number row
                continue

            buffer.append(
                EmployeeRecord(
                    batch_id=batch.id,
                    company_id=target_company_id,
                    created_by=current.id,
                    **record,
                )
            )
            total += 1
            if len(buffer) >= BATCH_SIZE:
                db.add_all(buffer)
                buffer.clear()

        if buffer:
            db.add_all(buffer)

        batch.total_records = total
        batch.processed_records = 0
        batch.status = BatchStatus.pending
        await db.commit()

    except Exception as exc:  # noqa: BLE001
        await db.rollback()
        batch.status = BatchStatus.error
        await db.commit()
        raise HTTPException(400, detail=f"Parse error: {exc}") from exc

    background_tasks.add_task(
        _log,
        db=db,
        user_id=current.id,
        entity_type="batch",
        entity_id=batch.id,
        action="upload_xlsx",
        details={"filename": file.filename, "records": batch.total_records},
    )
    return batch
